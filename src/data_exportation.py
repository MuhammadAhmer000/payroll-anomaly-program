import logging
import psycopg2

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from src.data_ingestion import load_payroll
import pandas as pd
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from src.database_ingestion import connect_database

logger = logging.getLogger(__name__)

def compute_zscore_output(writer, emp_id, rule_df, zscore_df, unsupervised_results, ensemble_df, root_cause_df_output, root_cause_summary_output):
    logger.info(f"Generating Excel output for employee {emp_id}")
    start_row = 1
    emp_id = str(emp_id)

    # --- Create the sheet manually and register it in writer.sheets ---
    sheet = writer.book.create_sheet(emp_id)
    writer.sheets[emp_id] = sheet
    worksheet = writer.sheets[emp_id]

    # --- Z-Score ---
    logger.debug("Writing z-score section")
    worksheet.cell(row=start_row, column=1, value="Z-Score Percentile")
    start_row += 1
    zscore_df.to_excel(writer, sheet_name=emp_id, startrow=start_row, index=False)
    start_row += len(zscore_df) + 3

    # --- Payroll Rule ---
    logger.debug("Writing payroll rule section")
    worksheet.cell(row=start_row, column=1, value="Payroll Rule Results")
    start_row += 1
    rule_df.to_excel(writer, sheet_name=emp_id, startrow=start_row, index=False)
    start_row += len(rule_df) + 3

    # --- Unsupervised ML Results ---
    logger.debug("Writing unsupervised ML section")
    worksheet.cell(row=start_row, column=1, value="Unsupervised Machine Learning Results")
    start_row += 1
    unsupervised_results.to_excel(writer, sheet_name=emp_id, startrow=start_row, index=False)
    start_row += len(unsupervised_results) + 3

    # --- Ensemble ML Summary ---
    logger.debug("Writing ensemble section")
    worksheet.cell(row=start_row, column=1, value="Ensemble ML Summary")
    start_row += 1
    ensemble_df.to_excel(writer, sheet_name=emp_id, startrow=start_row, index=False)
    start_row += len(ensemble_df) + 3

    # --- Root Cause Analysis ---
    logger.debug("Writing root cause section")
    worksheet.cell(row=start_row, column=1, value="Root Cause Analysis")
    start_row += 1
    root_cause_df_output.to_excel(writer, sheet_name=emp_id, startrow=start_row, index=False)
    start_row += len(root_cause_df_output) + 3

    # --- Root Cause Summary ---
    logger.debug("Writing root cause summary section")
    worksheet.cell(row=start_row, column=1, value="Root Cause Summary")
    start_row += 1
    root_cause_summary_output.to_excel(writer, sheet_name=emp_id, startrow=start_row, index=False)
    start_row += len(root_cause_summary_output) + 3

    # --- Overall Verdict ---
    logger.debug("Writing overall verdict section")
    worksheet.cell(row=start_row, column=1, value="Overall Verdict")
    start_row += 1
    overall_bool = (bool(zscore_df["anomaly"].any()) or ensemble_df["Ensemble_Anomaly"][0] or
                    ensemble_df["Ensemble_Anomaly"][0])
    overall_str = "Anomaly" if overall_bool else "No Anomaly"
    worksheet.cell(row=start_row, column=1, value=overall_str)
    start_row += 1

    logger.info(f"Excel output completed for employee {emp_id}")


from openpyxl import load_workbook, Workbook
from pathlib import Path


from openpyxl import load_workbook, Workbook


def output_database(results, database_credentials):

    with connect_database(*database_credentials) as conn:
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS Anomaly(
                    employee_id     NUMERIC(20,2),
                    zscore_cumulative           BOOLEAN,
                    payroll_rule_cumulative     BOOLEAN,
                    machine_learning_cumulative BOOLEAN,
                    overall_cumulative          BOOLEAN
                )
            """)

            rows = [
                (
                    result["emp_id"],
                    bool(result["stats_df"]["anomaly"].any()),
                    bool(result["rule_df"]["Anomaly"].any()),
                    bool(result["ensemble_df"]["Ensemble_Anomaly"][0]),
                    bool(result["overall"])
                )
                for result in results
            ]

            cur.executemany("""
                INSERT INTO Anomaly (
                    employee_id,
                    zscore_cumulative,
                    payroll_rule_cumulative,
                    machine_learning_cumulative,
                    overall_cumulative
                ) VALUES (%s, %s, %s, %s, %s)
            """, rows)





    print(conn)


# 2026-03-09 22:30:20 | DEBUG | src.sql_layer | Table query created for simulation: CREATE TABLE Payroll (month TEXT,
# employee_id NUMERIC(20,2), employee_name TEXT, department TEXT, designation TEXT, month_days NUMERIC(20,2), paid_days NUMERIC(20,2),
# lwp NUMERIC(20,2), ctc_annual NUMERIC(20,2), gross_salary NUMERIC(20,2), basic NUMERIC(20,2), hra NUMERIC(20,2),
# special_allowance NUMERIC(20,2), conveyance NUMERIC(20,2), medical NUMERIC(20,2), other_allowances NUMERIC(20,2),
# variable_pay NUMERIC(20,2), investments_80c NUMERIC(20,2), total_earnings NUMERIC(20,2), pf NUMERIC(20,2), pt NUMERIC(20,2),
# tds NUMERIC(20,2), other_deductions NUMERIC(20,2), total_deductions NUMERIC(20,2), net_payable NUMERIC(20,2));

# per numeric category, whether it deviated or not, "None" if it was dropped
# ISF, RISF, LOF, PCA, Ensemble_Verdict
# Overall Verdict



